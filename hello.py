from flask import Flask, request, jsonify, abort, redirect, url_for, render_template, send_file, flash
from bs4 import BeautifulSoup
import requests, statistics
import openpyxl
from flask_wtf import FlaskForm
from wtforms import StringField, FileField
from wtforms.validators import DataRequired
import os
from werkzeug.utils import secure_filename
import auto
import json

app = Flask(__name__, static_url_path='/static')

app.config.update(dict(
    SECRET_KEY="powerful secretkey",
    WTF_CSRF_SECRET_KEY="a csrf secret key"
))

    
@app.route('/')
def redir_submit():
    return redirect(url_for('index'))

class MyForm(FlaskForm):
    # name = 'name'
    file = FileField(validators=[DataRequired()])
# @app.route('/index', methods=['GET', 'POST'])
# def form():

#     #Причем в начале проверяем наличие авторизации, если флага нет, то кидаем обработку 401 ошибки и не даем работать с прогой
#     # if not session.get('logged_in'):
#     #     abort(401)

#     """
#     Тут делаем что-то полезное в случае успешной авторизации
#     """

#     return render_template('index.html')


@app.route('/submit', methods=('GET', 'POST'))
def submit():
    form = MyForm()
    
    if form.validate_on_submit():
        f = form.file.data
        filename = 'analogi.xlsx'
        f.save(os.path.join(filename))
        auto.autoru_appraiser(filename)
        

        return send_file(filename,
                     mimetype='xlsx',
                    #  attachment_filename=filename,
                     as_attachment=True)
        
    return render_template('submit.html', form=form)


@app.errorhandler(500)
def page_not_found(e):
    error = 'Произошла ошибка при работе скрипта, вероятно auto.ru опять показывает капчу. Сообщите об этом Алексею прямо сейчас.'
    return render_template('index.html', error=error), 500


def get_dropdown_values():

    """
    dummy function, replace with e.g. database call. If data not change, this function is not needed but dictionary
    could be defined globally
    """
    with open('marks_and_models.json', 'r', encoding='utf-8') as fh: #открываем файл на чтение
        class_entry_relations = json.load(fh)

    return class_entry_relations


@app.route('/_update_dropdown')
def update_dropdown():

    # the value of the first dropdown (selected by the user)
    selected_class = request.args.get('selected_class', type=str)

    # get values for the second dropdown
    updated_values = get_dropdown_values()[selected_class]

    # create the value sin the dropdown as a html string
    html_string_selected = ''
    for entry in updated_values:
        html_string_selected += '<option value="{}">{}</option>'.format(entry, entry)

    return jsonify(html_string_selected=html_string_selected)


@app.route('/_process_data')
def process_data():
    selected_class = request.args.get('selected_class', type=str)
    selected_entry = request.args.get('selected_entry', type=str)

    # process the two selected values here and return the response; here we just create a dummy string

    return jsonify(random_text="you selected {} and {}".format(selected_class, selected_entry))


@app.route('/index', methods=('GET', 'POST'))
def index():

    """
    Initialize the dropdown menues
    """
    form = MyForm()
    
    if request.method == 'POST':
        filename = 'shablon.xlsx'
        f = openpyxl.load_workbook(filename)
        sheetobject = f['Объекты оценки']
        sheetobject.cell(row=2, column=1).value = request.form['all_classes']
        sheetobject.cell(row=2, column=2).value = request.form['all_entries']
        sheetobject.cell(row=2, column=3).value = request.form['obj_year']
        sheetobject.cell(row=2, column=4).value = request.form['obj_engvol']
        sheetobject.cell(row=2, column=5).value = request.form['obj_hp']
        sheetobject.cell(row=2, column=9).value = request.form['obj_kpp']
        sheetobject.cell(row=2, column=7).value = request.form['obj_gear']
        sheetobject.cell(row=2, column=6).value = request.form['obj_eng']
        sheetobject.cell(row=2, column=8).value = request.form['obj_tob']
        sheetobject.cell(row=2, column=11).value = request.form['obj_mileage']
        filename = 'analogi.xlsx'
        
        f.save(os.path.join(filename))
        
        auto.autoru_appraiser(filename)
        

        return send_file(filename,
                     mimetype='xlsx',
                    #  attachment_filename=filename,
                     as_attachment=True)

    class_entry_relations = get_dropdown_values()

    default_classes = sorted(class_entry_relations.keys())
    default_values = class_entry_relations[default_classes[0]]

    return render_template('index.html',
                           all_classes=default_classes,
                           all_entries=default_values)

if __name__ == "__main__":
    app.run(debug=True)