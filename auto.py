from http import cookies
from os import lseek
from pprint import pprint
from bs4 import BeautifulSoup
from random import choice
import random
import requests
import openpyxl
from fake_useragent import UserAgent
import statistics

wbname = 'Аналоги.xlsx'
def autoru_appraiser(wbname):
    print(1)
    wb = openpyxl.load_workbook(wbname)
    sheetanalog = wb['Аналоги']
    sheetobject = wb['Объекты оценки']
    columns = ['Марка',	'Модель','Год выпуска','Объем двигателя','Мощность двигателя','Тип двигателя',	'Привод',	'Тип кузова',	'Тип КПП',	'Цвет',	'Пробег',	'Цена',	'Ссылка', '', 'Цена с корректировкой на торг', 'Цена с корректировкой на год', 'Цена с коррестировкой на пробег', 'Валовая корректировка', 'Стоимость с учетом всех корректировок']
    iobj = 2
    
    
    while sheetobject.cell(row=iobj, column=1).value != '':
        objmarka = sheetobject.cell(row=iobj, column=1).value
        objmodel = sheetobject.cell(row=iobj, column=2).value
        objyear = sheetobject.cell(row=iobj, column=3).value
        objengvol = sheetobject.cell(row=iobj, column=4).value
        objhp = sheetobject.cell(row=iobj, column=5).value
        objkpp = sheetobject.cell(row=iobj, column=9).value
        objgear = sheetobject.cell(row=iobj, column=7).value
        objeng = sheetobject.cell(row=iobj, column=6).value
        objtob = sheetobject.cell(row=iobj, column=8).value
        objmileage = sheetobject.cell(row=iobj, column=11).value
        
        try:
            objmarka = objmarka.replace('-', '_')
            objmarka = objmarka.replace(' ', '_')
        except:
            pass
        
        try:
            objmodel = objmodel.replace('-', '_')
            objmodel = objmodel.replace(' ', '_')
        except:
            pass
        
        #objeng = sheetobject.cell(row=iobj, column=6).value
        #КПП
        if objkpp == 'робот' or objkpp == 'автомат' or objkpp == 'вариатор':
            objkpp = 'AUTOMATIC&transmission=ROBOT&transmission=VARIATOR'
        elif objkpp == 'механика':
            objkpp = 'MECHANICAL'
        
        #ПРИВОД
        if objgear == 'полный':
            objgear = 'ALL_WHEEL_DRIVE'
        if objgear == 'задний':
            objgear = 'REAR_DRIVE'
        if objgear == 'передний':
            objgear = 'FORWARD_CONTROL'
        #Тип двигателя
        if objeng == 'бензин':
            objeng = 'GASOLINE'
        elif objeng == 'дизель':
            objeng = 'DIESEL'
        elif objeng == 'гибрид':
            objeng = 'HYBRID'
        elif objeng == 'электро':
            objeng = 'ELECTRO'
        #Тип кузова
        if objtob == 'седан':
            objtob = 'SEDAN'
        elif objtob == 'хэтчбек':
            objtob = 'HATCHBACK'
        elif objtob == 'хэтчбек 3дв.':
            objtob = 'hatchback_3_doors'
        elif objtob == 'хэтчбек 5дв.':
            objtob = 'hatchback_5_doors'
        elif objtob == 'внедорожник':
            objtob = 'allroad'
        elif objtob == 'внедорожник 3дв.':
            objtob = 'allroad_3_doors'
        elif objtob == 'внедорожник 5дв.':
            objtob = 'allroad_5_doors'
        elif objtob == 'универсал':
            objtob = 'wagon'
        elif objtob == 'купе':
            objtob = 'coupe'
        elif objtob == 'минивэн':
            objtob = 'minivan'
        elif objtob == 'пикап':
            objtob = 'pickup'
        elif objtob == 'лимузин':
            objtob = 'limousine'
        elif objtob == 'фургон':
            objtob = 'van'
        elif objtob == 'кабриолет':
            objtob = 'cabrio'
        #objgear = sheetobject.cell(row=iobj, column=7).value
        #objtob = sheetobject.cell(row=iobj, column=8).value
        #objkpp = sheetobject.cell(row=iobj, column=9).value
        
        #print(objyear)
        try:
            year1 = int(objyear) - 1
            
        except:
            break
        print(year1, objyear)
        year2 = int(objyear) + 1

        hp1 = round(float(objhp) * 0.95)
        hp2 = round(float(objhp) * 1.05)

        vol1 = round(int(objengvol) * 0.9,-2)
        if 3000 < vol1 < 3500:
            vol1 = 3000
        if 3500 < vol1 < 4000:
            vol1 = 3500
        if 4000 < vol1 < 4500:
            vol1 = 4000
        if 4500 < vol1 < 5000:
            vol1 = 4500
        if 5000 < vol1 < 5500:
            vol1 = 5000
        if 5500 < vol1 < 6000:
            vol1 = 5500
        if 6000 < vol1 < 7000:
            vol1 = 6000
        if 7000 < vol1 < 8000:
            vol1 = 7000
        if 8000 < vol1 < 9000:
            vol1 = 8000
        if 9000 < vol1:
            vol1 = 9000
        vol2 = round(int(objengvol) * 1.1,-2)
        if 3000 < vol2 < 3500:
            vol2 = 3500
        if 3500 < vol2 < 4000:
            vol2 = 4000
        if 4000 < vol2 < 4500:
            vol2 = 4500
        if 4500 < vol2 < 5000:
            vol2 = 5000
        if 5000 < vol2 < 5500:
            vol2 = 5500
        if 5500 < vol2 < 6000:
            vol2 = 6000
        if 6000 < vol2 < 7000:
            vol2 = 7000
        if 7000 < vol2 < 8000:
            vol2 = 8000
        if 8000 < vol2 < 9000:
            vol2 = 9000
        if 9000 < vol2 < 10000:
            vol2 = 10000

        millage1 = int(objmileage) * 0.50
        millage2 = int(objmileage) * 1.50
        if millage2 < 10000:
            millage2 = millage2 + 30000
        if millage2 > 200000:
            millage1 = 100000
        totalprice = 0
        totalpriceall = 0
        totalpricemedian = []
        
        url = 'https://auto.ru/cars/' + str(objmarka) + '/' + str(objmodel) + '/all/body-' + str(objtob) + '/?year_from=' + str(year1) + '&year_to=' + str(year2) + '&power_from=' + str(hp1) + '&displacement_from=' + str(vol1).replace('.0','') + '&displacement_to=' + str(vol2).replace('.0','') + '&transmission=' + str(objkpp) + '&power_to=' + str(hp2) + '&km_age_from=' + str(round(millage1,-1)).replace('.0','') + '&km_age_to=' + str(round(millage2,-1)).replace('.0','') + '&engine_group=' + str(objeng) + '&gear_type=' + str(objgear)
        #body_type_group = SEDAN & body_type_group = HATCHBACK & body_type_group = HATCHBACK_3_DOORS & body_type_group = HATCHBACK_5_DOORS & body_type_group = LIFTBACK & body_type_group = ALLROAD & body_type_group = ALLROAD_3_DOORS & body_type_group = ALLROAD_5_DOORS & body_type_group = WAGON & body_type_group = COUPE & body_type_group = MINIVAN & body_type_group = PICKUP & body_type_group = LIMOUSINE & body_type_group = VAN & body_type_group = CABRIO
        #ROBOT&transmission=AUTOMATIC&transmission=VARIATOR&transmission=AUTO
        #GASOLINE&engine_group=DIESEL&engine_group=HYBRID&engine_group=ELECTRO
        #FORWARD_CONTROL&gear_type=REAR_DRIVE&gear_type=ALL_WHEEL_DRIVE
        print(url)
                
        user = UserAgent().random
        cookie_one = {'_csrf_token': '4339396c17d0de9eae9f1c650ec6066a45eab30e862fd11a', 
                'suid': 'e43436fa3922a09ece744acad847503e.8fd345446dfa3dae2686e14b002e5a12',
                'from': 'direct', 
                'gdpr': '0',
                '_ym_uid': '1637562172583098176', 
                'BCSI-CS-e9746730d4af8182': '2', 
                'deal_million_popup_page_seen': '-1', 
                'gids': '', 
                'yandex_login': '', 
                'i': '8G6VlKR8YEZvnjkNFTIFPZe2rF2H5KWi69dPC09SbC9CJQNViEIvo5s06pT+nQtJZj+lhsfJyJ5xGS/TT8826ybrqDU=', 
                'gradius': '200', 
                'BCSI-CS-31c2a4b3bbb1d3f2': '2', 
                'yandexuid': '3780874931628665890', 
                'my': 'YwA%3D', 
                'autoruuid': 'g62ac378b23hvln2euamj42uhve65tvs.95f9ed57c37bcb7f519e00b7629f14f4', 
                'safe_deal_promo': '3', 
                'panorama_press_and_spin_closed': 'true', 
                'yuidlt': '1', 
                'autoru_sid': 'a%3Ag62ac378b23hvln2euamj42uhve65tvs.95f9ed57c37bcb7f519e00b7629f14f4%7C1656335449507.604800.sj70OMADbmrSXwek-p4Srw.2bX02SYAKC-p67hhzvcU0PBvs-lBCTApLSn8ATfEWNg', 
                'spravka': 'dD0xNjU2MzQwMDE4O2k9OTEuMjIzLjYzLjcxO0Q9M0I0QjVFNUFFNjJGMjIwOEE5NjE0RENGOTM0REQ1OEJCRUUwMkNEQzE0Q0E0NjcxN0YxNDBEOTFGQjYyMkFDNTIzRjREOTUxO3U9MTY1NjM0MDAxODc5MzY2OTkxMTtoPTk0ZjUyMWM2MDFlNTA4ZDNhYTBlMjk5NWFmODY5NDJh', 
                '_yasc': '7yvXfv3UnghJNLlTdbNzKxvUIVXpMsokVAW6ksN2wxXuqaEf', 
                'autoru-visits-count': '9', 
                '_ym_isad': '2', 
                'ys': 'c_chck.3966782579', 
                'mda2_beacon': '1656572550206', 
                'from_lifetime': '1656572551611', 
                '_ym_d': '1656572551', 
                'cycada': 'vSZ8GwepZ9WvKrK76MQTehBqOB75p8DO44Pj0lAPZes='}
        cookie_two = {'suid': '65834b02185b6cd61a4bc76cdc9a4fe7.5b97d40aac67cf9a9e6f3f0b29ae2738', 
                'gdpr': '0', 
                '_ym_uid': '1654251044667232301', 
                'yandex_login': 'dema.lex', 
                'i': 'aGjQyNZ/ZRO65SUHj4wbP23Z6tJI4SULlDWv1aObZTJOmWNWij30yDKW3BV75P16qIHY79JnfxHmSck2fZGOAPTDbXc=', 
                'gradius': '1100', 
                'gids': '', 
                'yandexuid': '9671052191649090137', 
                'autoru-visits-count': '1', 
                'yuidlt': '1', 
                'autoru_sid': 'a%3Ag6310cc6226vmijk6sb8d7rm5q1rbchr.db3207eb5930dd7472c188ac8831b3ea%7C1662987638286.604800.Upm04ekihi6YPeDzDsHz9g.6O-jtA7rWalispYFDBWutuy0rjSRmv2m4-HaHRFNAbY', 
                'crookie': 'CDc9bYHo+AX1Ap5MOeBlzGPXQVvFRiXtog37ZgqZjinA8NMwCCl/QCR7qGvop5PpSiMhlmTZ7wtN10geQdwRiDie9VA=', 
                'cmtchd': 'MTY2Mjk4NzY0MDE1OA==', 
                'autoruuid': 'g6310cc6226vmijk6sb8d7rm5q1rbchr.db3207eb5930dd7472c188ac8831b3ea', 
                '_ym_isad': '2', 
                '_ym_visorc': 'b', 
                'spravka': 'dD0xNjYzMzQwMjQ4O2k9MzEuMTczLjgwLjExNDtEPTRDNTFENkY2NzZBMTlGOUYyODI2NjJEQjhDMjhCMjY5Nzc5MkRENUJCMDU1RkY1ODZGRTkwMzlGOUQzRkZCRjY4QzdGNzg3Qzt1PTE2NjMzNDAyNDg1MzA4NDQzOTU7aD0yMzUwODE1M2QyODFiOGYyZWRhN2ViMmQyMTdlNGFlZQ==', 
                'Session_id': '3:1663340257.5.0.1649090824573:FEqwsg:20.1.2:1|401706160.0.2|61:10007439.183766.mLkUfhobq7_dQEjGS6D9QYOdemM', 
                'mda2_beacon': '1663340257300', 
                'sso_status': 'sso.passport.yandex.ru:synchronized', 
                '_csrf_token': 'c6f92559c01351f07eded71e07a2e8d5e73f627000cfb787', 
                'from': 'direct', 
                '_yasc': 'mfT9HR3PWBnnfO6xHtABm4Nmf9H0in1mpH/fhATTYe32kg6T', 
                'layout-config': '{"win_width":498.3999938964844,"win_height":736.7999877929688}', 
                '_ym_d': '1663340420', 
                'from_lifetime': '1663340420414', 
                'cycada': 'LX7X2UW1ZDYfOT4Erpc14yi3/gL9kZfLyu8xzELZaHA=' }
        cookies_list = [cookie_one, cookie_two]
        cookies = random.choice(cookies_list)
        headers = {'user-agent': user,
                "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9", 
                "Accept-Encoding": "gzip, deflate, br", 
                "Accept-Language": "ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7", 
                "Sec-Fetch-Dest": "document", 
                "Sec-Fetch-Mode": "navigate", 
                "Sec-Fetch-Site": "cross-site", 
                "Sec-Fetch-User": "?1", 
                "Upgrade-Insecure-Requests": "1", }#'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/103.0.0.0 Safari/537.36'}
        s = requests.Session()
        resp = s.get('https://auto.ru/', headers=headers, cookies=cookies)#, headers=headers
        resp.encoding = 'utf-8'
        if 'Ой!' in resp.text:
            print(resp.request.headers)
            print(s.headers)
            print('first fuck!!!!!')
            # driver.quit()
            break
        
        print(s.cookies)
        resp = s.get(url, headers=headers, cookies=cookies)
        print(s.headers)
        resp.encoding = 'utf-8'
        print('5')
        soup = BeautifulSoup(resp.text, 'html.parser')
        if 'Ничего не найдено' in soup.text or 'Страница не найдена' in soup.text:
            sheetobject.cell(row=iobj, column=12).value = 'Аналоги не найдены. Попробуйте изменить параметры ТС'
            iobj = iobj + 1
            # driver.quit()
            continue
        if 'Ой!' in soup.text:
            print('fuck!!!!!')
            # driver.quit()
            break
        #print(soup.text)
        
        marka = soup.findAll('a', {'class' : 'Link BreadcrumbsGroup__itemText'})
        marka = marka[1].text
        model = soup.find('span', {'class' : 'BreadcrumbsGroup__itemText'}).text.replace(marka+' ','')
        car_list = soup.findAll('div', {'class' : 'ListingItem'})
        shname = 'Аналоги ' + str(iobj-1)
        try:
            delete = wb[shname]
            wb.remove(delete)
            wsanalog = wb.create_sheet(shname)
        except:
            wsanalog = wb.create_sheet(shname)
        #wsanalog = wb.create_sheet(shname)
        for iws, value in enumerate(columns, 1):
            wsanalog.cell(row=1, column=iws).value = value
        i = 2
        
        for car in car_list:
            try:
                carname = car.find('a', {'class' : 'Link ListingItemTitle__link'}).text
                caryear = car.find('div', {'class' : 'ListingItem__year'}).text
                cartech = car.findAll('div', {'class' : 'ListingItemTechSummaryDesktop__cell'})
                carlink = car.find('a', {'class' : 'Link OfferThumb'}).get('href')
                carmileage = car.find('div', {'class' : 'ListingItem__kmAge'}).text.replace(' км', '').replace(' ','')
                carkpp = cartech[1].text
                cartob = cartech[2].text
                cargear = cartech[3].text
                cartech_spl = cartech[0].text.split('/')
                carengvol = cartech_spl[0].replace(' л', '')
                carenghp = cartech_spl[1].replace(' л.с.', '')
                careng = cartech_spl[2].lower().replace(' ', '')
                carcolor = cartech[4].text
                carprice = car.findAll('span')
                try:
                    carprice = int(carprice[0].text.replace(' ₽','').replace(' ',''))
                except:
                    carprice = car.find('div', {'class' : 'ListingItemPrice__content'}).text
                    carprice = int(carprice.replace(' ₽', '').replace(' ', ''))
                wsanalog.cell(row=i, column=1).value = marka
                wsanalog.cell(row=i, column=2).value = model
                wsanalog.cell(row=i, column=3).value = int(caryear)
                wsanalog.cell(row=i, column=4).value = float(carengvol)
                wsanalog.cell(row=i, column=5).value = int(carenghp)
                wsanalog.cell(row=i, column=6).value = careng
                wsanalog.cell(row=i, column=7).value = cargear
                wsanalog.cell(row=i, column=8).value = cartob
                wsanalog.cell(row=i, column=9).value = carkpp
                wsanalog.cell(row=i, column=10).value = carcolor
                wsanalog.cell(row=i, column=11).value = int(carmileage)
                wsanalog.cell(row=i, column=12).value = int(carprice)
                wsanalog.cell(row=i, column=13).value = carlink
                wsanalog.cell(row=i, column=15).value = torgprice = int(int(carprice)*0.9)
                if int(objyear) == int(caryear):
                    wsanalog.cell(row=i, column=16).value = yearprice = int(torgprice)
                    yearcor = 0
                elif int(caryear) == int(objyear)-1:
                    wsanalog.cell(row=i, column=16).value = yearprice = int(int(torgprice)*1.07)
                    yearcor = -7
                elif int(caryear) == int(objyear)+1:
                    wsanalog.cell(row=i, column=16).value = yearprice = int(int(torgprice)*0.93)
                    yearcor = 7
                if int(objmileage)*0.95 < int(carmileage) <= int(objmileage)*0.98:
                    wsanalog.cell(row=i, column=17).value = millageprice = int(int(yearprice) * 0.99)
                    millagecor = -1
                elif int(objmileage)*0.98 < int(carmileage) < int(objmileage)*1.02:
                    wsanalog.cell(row=i, column=17).value = millageprice = int(yearprice)
                    millagecor = 0
                elif int(objmileage)*0.93 < int(carmileage) <= int(objmileage)*0.95:
                    wsanalog.cell(row=i, column=17).value = millageprice = int(int(yearprice)*0.98)
                    millagecor = -2
                elif int(objmileage)*0.9 < int(carmileage) <= int(objmileage)*0.93:
                    wsanalog.cell(row=i, column=17).value = millageprice = int(int(yearprice)*0.97)
                    millagecor = -3
                elif int(objmileage)*0.87 < int(carmileage) <= int(objmileage)*0.9:
                    wsanalog.cell(row=i, column=17).value = millageprice = int(int(yearprice)*0.96)
                    millagecor = -4
                elif int(objmileage)*0.83 < int(carmileage) <= int(objmileage)*0.87:
                    wsanalog.cell(row=i, column=17).value = millageprice = int(int(yearprice)*0.95)
                    millagecor = -5
                elif int(objmileage)*0.80 < int(carmileage) <= int(objmileage)*0.83:
                    wsanalog.cell(row=i, column=17).value = millageprice = int(int(yearprice)*0.94)
                    millagecor = -6
                elif int(carmileage) <= int(objmileage)*0.80:
                    wsanalog.cell(row=i, column=17).value = millageprice = int(int(yearprice)*0.98)
                    millagecor = -7
                elif int(objmileage)*1.02 <= int(carmileage) < int(objmileage)*1.05:
                    wsanalog.cell(row=i, column=17).value = millageprice = int(int(yearprice) * 1.01)
                    millagecor = 1
                elif int(objmileage)*1.05 <= int(carmileage) < int(objmileage)*1.07:
                    wsanalog.cell(row=i, column=17).value = millageprice = int(int(yearprice)*1.02)
                    millagecor = 2
                elif int(objmileage)*1.07 <= int(carmileage) < int(objmileage)*1.1:
                    wsanalog.cell(row=i, column=17).value = millageprice = int(int(yearprice)*1.03)
                    millagecor = 3
                elif int(objmileage)*1.1 <= int(carmileage) < int(objmileage)*1.13:
                    wsanalog.cell(row=i, column=17).value = millageprice = int(int(yearprice)*1.04)
                    millagecor = 4
                elif int(objmileage)*1.13 <= int(carmileage) < int(objmileage)*1.17:
                    wsanalog.cell(row=i, column=17).value = millageprice = int(int(yearprice)*1.05)
                    millagecor = 5
                elif int(objmileage)*1.17 <= int(carmileage) < int(objmileage)*1.20:
                    wsanalog.cell(row=i, column=17).value = millageprice = int(int(yearprice)*1.06)
                    millagecor = 6
                elif int(objmileage)*1.20 <= int(carmileage):
                    wsanalog.cell(row=i, column=17).value = millageprice = int(int(yearprice)*1.07)
                    millagecor = 7
                wsanalog.cell(row=i, column=18).value = valcor = yearcor + millagecor
                wsanalog.cell(row=i, column=19).value = round(millageprice,-3)

                #print(i,year1, objyear)

                totalpriceall = totalpriceall + round(millageprice,-3)
                totalpricemedian.append(round(millageprice,-3))
                i = i + 1
                wb.save(wbname)
            except:
                pass
        wsanalog.cell(row=1, column=20).value = url
        wb.save(wbname)
        
        print('Отсев')
        ip = 2
        ic = i
        
        print(statistics.median(totalpricemedian))
        totalprice = round(totalpriceall / (ic - 2), -3)
        print(totalprice)
        while ip <= ic-1:
            #ПРИВОД
            #print(ip, 'привод', wsanalog.cell(row=ip, column=7).value.lower())
            if sheetobject.cell(row=iobj, column=7).value != wsanalog.cell(row=ip, column=7).value or sheetobject.cell(row=iobj, column=6).value != wsanalog.cell(row=ip, column=6).value.replace(' ', '') or hp1 > wsanalog.cell(row=ip, column=5).value or wsanalog.cell(row=ip, column=5).value > hp2:
                totalpriceall = totalpriceall - int(wsanalog.cell(row=ip, column=19).value)
                totalpricemedian.remove(int(wsanalog.cell(row=ip, column=19).value))
                wsanalog.delete_rows(ip, 1)
                print('удаляем', ip)
                wb.save(wbname)
                ic=ic-1
                ip = ip - 1
                        
            ip = ip + 1
            if wsanalog.cell(row=ip-1, column=19).value == '':
                break
        ip = 2
        try:
            totalprice = round(totalpriceall / (ic - 2), -3)
            #print(statistics.median(totalpricemedian))
            #print(totalprice)
            while ip <= ic-1:
                if int(wsanalog.cell(row=ip, column=19).value)/statistics.median(totalpricemedian) > 1.15 or int(wsanalog.cell(row=ip, column=19).value)/statistics.median(totalpricemedian) < 0.85:
                    totalpriceall = totalpriceall - int(wsanalog.cell(row=ip, column=19).value)
                    totalpricemedian.remove(int(wsanalog.cell(row=ip, column=19).value))
                    wsanalog.delete_rows(ip, 1)
                    wb.save(wbname)
                    ic=ic-1
                    totalprice = round(totalpriceall / (ic-2), -3)
                    print('удаляем', ip)
                    ip = ip - 1
                ip = ip + 1
                if wsanalog.cell(row=ip-1, column=19).value == '':
                    break
            ip = 2
            #ic = i
            
            totalprice = round(totalpriceall / (ic - 2), -3)
            #print(statistics.median(totalpricemedian))
            #print(totalprice)
            while ip <= ic - 1:
                if int(wsanalog.cell(row=ip, column=19).value) / statistics.median(totalpricemedian) > 1.17 or int(
                        wsanalog.cell(row=ip, column=19).value) / statistics.median(totalpricemedian) < 0.83:
                    totalpriceall = totalpriceall - int(wsanalog.cell(row=ip, column=19).value)
                    wsanalog.delete_rows(ip, 1)
                    wb.save(wbname)
                    ic = ic - 1
                    totalprice = round(totalpriceall / (ic - 2), -3)
                    print('удаляем', ip)
                    ip = ip - 1
                ip = ip + 1
                if wsanalog.cell(row=ip - 1, column=19).value == '':
                    break
            print(iobj-1, year1, totalprice, statistics.median(totalpricemedian))
            #totalprice = round(totalpriceall/(iс-2),-3)
        except:
            totalprice = 'Аналоги не найдены. Попробуйте изменить параметры ТС'

        sheetobject.cell(row=iobj, column=12).value = totalprice
        if sheetobject.cell(row=iobj, column=1).value == '':
            break
        # driver.quit()
        iobj = iobj + 1

    wb.save(wbname)